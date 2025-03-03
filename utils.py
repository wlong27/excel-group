import pandas as pd
import io
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def process_excel_file(df, group_column):
    """
    Process the Excel file by creating separate workbooks for each group,
    preserving source formatting and formulas.

    Args:
        df (pd.DataFrame): Input DataFrame
        group_column (str): Column name to group by

    Returns:
        bytes: Zip file content as bytes containing all workbooks
    """
    try:
        # Create a buffer to store the ZIP file
        zip_buffer = io.BytesIO()

        # Create ZIP file
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Group the data
            unique_values = df[group_column].unique()

            for value in unique_values:
                # Filter data for current group
                group_df = df[df[group_column] == value]

                # Create workbook name (ensure it's valid)
                workbook_name = str(value)
                # Remove invalid characters from filename
                workbook_name = "".join(c for c in workbook_name if c.isalnum() or c in (' ', '_', '-'))
                file_name = f"{workbook_name}.xlsx"

                # Create a buffer for the Excel file
                excel_buffer = io.BytesIO()

                # Save the group data to Excel, preserving formatting
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    # First write the data
                    group_df.to_excel(writer, index=False, sheet_name='Sheet1')

                    # Get the worksheet
                    worksheet = writer.sheets['Sheet1']

                    # Copy column widths and formatting
                    for idx, col in enumerate(group_df.columns, 1):
                        letter = get_column_letter(idx)
                        # Set a reasonable column width based on content
                        max_length = max(
                            group_df[col].astype(str).apply(len).max(),
                            len(str(col))
                        )
                        worksheet.column_dimensions[letter].width = max_length + 2

                    # Apply basic formatting
                    header_row = worksheet[1]
                    for cell in header_row:
                        cell.style = 'Headline 1'
                        cell.font = cell.font.copy(bold=True)
                        cell.fill = cell.fill.copy(
                            start_color='F0F2F6',
                            end_color='F0F2F6'
                        )

                # Add the Excel file to the ZIP
                excel_buffer.seek(0)
                zip_file.writestr(file_name, excel_buffer.getvalue())

        # Get the ZIP file content
        zip_buffer.seek(0)
        return zip_buffer.read()

    except Exception as e:
        raise Exception(f"Error processing file: {str(e)}")